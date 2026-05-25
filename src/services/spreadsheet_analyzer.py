import logging
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    import xlrd  # type: ignore
except Exception:  # pragma: no cover
    xlrd = None


class SpreadsheetAnalyzer:
    MAX_ROWS_PER_SHEET = 5000

    def __init__(self) -> None:
        self.logger = logging.getLogger(__name__)

    def normalize_header(self, name: str) -> str:
        text = unicodedata.normalize("NFKD", str(name or ""))
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = text.lower().strip()
        text = re.sub(r"\s+", " ", text)
        return text

    def detect_headers(self, rows: list[list[Any]]) -> list[str]:
        for row in rows[:20]:
            values = ["" if v is None else str(v).strip() for v in row]
            non_empty = [v for v in values if v]
            if len(non_empty) >= 2:
                return [v or f"col_{idx + 1}" for idx, v in enumerate(values)]
        return []

    def load_workbook_tables(self, path: str) -> list[dict]:
        extension = Path(path).suffix.lower()
        if extension == ".xlsx":
            return self._load_xlsx_tables(path)
        if extension == ".xls":
            return self._load_xls_tables(path)
        raise ValueError(f"Formato de hoja de cálculo no soportado: {extension}")

    def _load_xlsx_tables(self, path: str) -> list[dict]:
        wb = load_workbook(path, read_only=True, data_only=True)
        tables: list[dict] = []
        for sheet in wb.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else value for value in row]
                if any(str(v).strip() for v in values):
                    rows.append(values)
                if len(rows) >= self.MAX_ROWS_PER_SHEET:
                    break
            table = self._build_table(sheet.title, rows)
            if table:
                tables.append(table)
        self.logger.info("SpreadsheetAnalyzer: hojas xlsx detectadas=%s", len(tables))
        return tables

    def _load_xls_tables(self, path: str) -> list[dict]:
        if xlrd is None:
            raise RuntimeError("xlrd no está instalado para procesar .xls")
        book = xlrd.open_workbook(path, on_demand=True)
        tables: list[dict] = []
        for sheet in book.sheets():
            rows = []
            for row_idx in range(min(sheet.nrows, self.MAX_ROWS_PER_SHEET)):
                values = sheet.row_values(row_idx)
                if any(str(v).strip() for v in values):
                    rows.append(values)
            table = self._build_table(sheet.name, rows)
            if table:
                tables.append(table)
        self.logger.info("SpreadsheetAnalyzer: hojas xls detectadas=%s", len(tables))
        return tables

    def _build_table(self, sheet_name: str, rows: list[list[Any]]) -> dict | None:
        if not rows:
            return None
        headers = self.detect_headers(rows)
        if not headers:
            return None
        header_len = len(headers)
        data_rows: list[dict] = []
        for row in rows[1:]:
            padded = list(row) + [""] * max(0, header_len - len(row))
            row_dict = {headers[idx]: padded[idx] for idx in range(header_len)}
            if any(str(v).strip() for v in row_dict.values()):
                data_rows.append(row_dict)
        self.logger.info(
            "SpreadsheetAnalyzer: hoja=%s cabeceras=%s filas=%s",
            sheet_name,
            len(headers),
            len(data_rows),
        )
        return {"sheet": sheet_name, "headers": headers, "rows": data_rows}

    def aggregate_by_column(self, path: str, group_column: str, value_column: str) -> dict[str, float]:
        tables = self.load_workbook_tables(path)
        group_norm = self.normalize_header(group_column)
        value_norm = self.normalize_header(value_column)
        result: defaultdict[str, float] = defaultdict(float)
        rows_processed = 0

        for table in tables:
            header_map = {self.normalize_header(h): h for h in table["headers"]}
            group_header = self._match_header(header_map, group_norm)
            value_header = self._match_header(header_map, value_norm)
            if not group_header or not value_header:
                continue

            self.logger.info(
                "SpreadsheetAnalyzer: agrupando hoja=%s group_col=%s value_col=%s",
                table["sheet"],
                group_header,
                value_header,
            )
            for row in table["rows"]:
                group_value = str(row.get(group_header, "")).strip()
                raw_value = row.get(value_header, "")
                numeric_value = self._to_float(raw_value)
                if not group_value:
                    continue
                result[group_value] += numeric_value
                rows_processed += 1

        self.logger.info(
            "SpreadsheetAnalyzer: filas_procesadas=%s grupos=%s",
            rows_processed,
            len(result),
        )
        return dict(sorted(result.items(), key=lambda item: item[0].lower()))

    def _match_header(self, header_map: dict[str, str], target: str) -> str | None:
        aliases = {
            "neto": ["neto", "kg neto", "kilos neto", "kilogramos neto"],
            "variedad": ["variedad", "tipo variedad"],
        }
        candidates = aliases.get(target, [target])
        for candidate in candidates:
            if candidate in header_map:
                return header_map[candidate]
        for normalized, original in header_map.items():
            if any(candidate in normalized for candidate in candidates):
                return original
        return None

    def _to_float(self, value: Any) -> float:
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value or "").strip().replace(" ", "")
        if not text:
            return 0.0
        text = text.replace(".", "").replace(",", ".") if text.count(",") == 1 and text.count(".") >= 1 else text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0
