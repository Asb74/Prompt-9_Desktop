import csv
import logging
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    import xlrd  # type: ignore
except Exception:  # pragma: no cover
    xlrd = None


class TableLoader:
    HEADER_KEYWORDS = {"socio", "nombre", "variedad", "neto", "importe", "precio"}

    def __init__(self) -> None:
        self.logger = logging.getLogger(__name__)

    def load_tables(self, path: str) -> list[dict[str, Any]]:
        extension = Path(path).suffix.lower()
        self.logger.info("TableLoader: archivo=%s tipo=%s", Path(path).name, extension)

        if extension == ".xlsx":
            return self._load_xlsx(path)
        if extension == ".xls":
            return self._load_xls(path)
        if extension == ".csv":
            return self._load_csv(path)

        raise ValueError(f"Formato tabular no soportado: {extension}")

    def detect_header_row(self, rows: list[list[Any]]) -> int:
        best_idx: int | None = None
        best_score = float("-inf")
        first_non_empty: int | None = None

        for idx, row in enumerate(rows[:20]):
            values = [self._cell_to_text(value) for value in row]
            non_empty = [value for value in values if value]
            if non_empty and first_non_empty is None:
                first_non_empty = idx
            if len(non_empty) < 2:
                continue

            text_count = sum(1 for value in non_empty if not self._looks_numeric(value))
            numeric_count = len(non_empty) - text_count
            score = (len(non_empty) * 3) + (text_count * 2) - (numeric_count * 2)

            if score > best_score:
                best_score = score
                best_idx = idx

        if best_idx is not None:
            return best_idx
        if first_non_empty is not None:
            return first_non_empty
        return 0

    def clean_header(self, value: Any, fallback_idx: int, used: dict[str, int]) -> str:
        raw = self._cell_to_text(value).replace("\n", " ").replace("\r", " ").strip()
        base = raw if raw else f"Columna_{fallback_idx + 1}"
        count = used.get(base, 0) + 1
        used[base] = count
        return base if count == 1 else f"{base}_{count}"

    def normalize_cell_value(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return value
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, time):
            return value.isoformat()

        text = self._cell_to_text(value).strip()
        if not text:
            return None

        parsed_number = self._parse_numeric_string(text)
        if parsed_number is not None:
            return parsed_number
        return text

    def table_to_preview_text(self, table: dict[str, Any], max_rows: int = 20) -> str:
        lines = [f"[Hoja: {table.get('sheet_name', 'SinNombre')}]", "Columnas detectadas:"]
        headers = table.get("headers", [])
        lines.append(" | ".join(str(header) for header in headers))
        lines.append("")
        lines.append("Filas:")

        for row in table.get("rows", [])[:max_rows]:
            pairs = [f"{header}={row.get(header)}" for header in headers]
            lines.append(" | ".join(pairs))

        return "\n".join(lines)

    def _load_xlsx(self, path: str) -> list[dict[str, Any]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        tables: list[dict[str, Any]] = []

        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                continue

            raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            table = self._build_table(sheet.title, raw_rows)
            if table is not None:
                tables.append(table)

        self.logger.info("TableLoader xlsx: hojas_leidas=%s", len(tables))
        return tables

    def _load_xls(self, path: str) -> list[dict[str, Any]]:
        if xlrd is None:
            raise RuntimeError("xlrd no está instalado para procesar .xls")

        workbook = xlrd.open_workbook(path, on_demand=True)
        tables: list[dict[str, Any]] = []

        for sheet in workbook.sheets():
            raw_rows = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
            table = self._build_table(sheet.name, raw_rows)
            if table is not None:
                tables.append(table)

        self.logger.info("TableLoader xls: hojas_leidas=%s", len(tables))
        return tables

    def _load_csv(self, path: str) -> list[dict[str, Any]]:
        parsed_rows: list[list[Any]] | None = None
        delimiters = [None, ";", ","]

        for encoding in ("utf-8-sig", "latin-1"):
            for forced_delimiter in delimiters:
                try:
                    parsed_rows = self._read_csv(path, encoding, forced_delimiter)
                    self.logger.info(
                        "TableLoader csv: encoding=%s delimiter=%s filas=%s",
                        encoding,
                        forced_delimiter or "sniffer",
                        len(parsed_rows),
                    )
                    break
                except Exception as exc:  # pragma: no cover
                    self.logger.warning(
                        "TableLoader csv: fallo lectura archivo=%s encoding=%s delimiter=%s error=%s",
                        Path(path).name,
                        encoding,
                        forced_delimiter or "sniffer",
                        exc,
                    )
            if parsed_rows is not None:
                break

        if parsed_rows is None:
            raise ValueError(f"No se pudo leer el CSV: {path}")

        table = self._build_table("CSV", parsed_rows)
        tables = [table] if table else []
        self.logger.info("TableLoader csv: tablas_detectadas=%s", len(tables))
        return tables

    def _build_table(self, sheet_name: str, rows: list[list[Any]]) -> dict[str, Any] | None:
        if not rows:
            self.logger.info("TableLoader: hoja_vacia=%s", sheet_name)
            return None

        blocks = self._detect_blocks(rows)
        if not blocks:
            self.logger.info("TableLoader: hoja_sin_bloques=%s", sheet_name)
            return None

        main_block = blocks[0]
        merged_rows = list(main_block["rows"])
        for block in blocks[1:]:
            if self._headers_compatible(main_block["headers"], block["headers"]):
                merged_rows.extend(self._remap_rows(block["rows"], block["headers"], main_block["headers"]))

        table = {
            "sheet_name": sheet_name,
            "headers": main_block["headers"],
            "rows": merged_rows,
            "row_count": len(merged_rows),
            "column_count": len(main_block["headers"]),
        }

        total_neto = self._sum_numeric_column(merged_rows, main_block["headers"], "neto")
        self.logger.info(
            "TableLoader diagnóstico: hoja=%s cabeceras_detectadas=%s filas_cabecera=%s bloques_detectados=%s filas_utiles_por_bloque=%s variedades_por_bloque=%s total_filas_procesadas=%s total_neto_calculado=%s",
            sheet_name,
            [b["headers"] for b in blocks],
            [b["header_row_index"] for b in blocks],
            len(blocks),
            [len(b["rows"]) for b in blocks],
            [sorted(list(b["varieties"])) for b in blocks],
            len(merged_rows),
            total_neto,
        )
        return table

    def _detect_blocks(self, rows: list[list[Any]]) -> list[dict[str, Any]]:
        blocks: list[dict[str, Any]] = []
        current: dict[str, Any] | None = None

        for row_idx, row in enumerate(rows):
            if not self._row_has_content(row):
                continue

            if self._is_header_candidate(row):
                if current and current["rows"]:
                    blocks.append(current)
                headers = self._build_headers_from_row(row)
                current = {
                    "headers": headers,
                    "header_row_index": row_idx,
                    "rows": [],
                    "varieties": set(),
                }
                continue

            if current is None:
                continue

            normalized_row = list(row)
            if len(normalized_row) < len(current["headers"]):
                normalized_row.extend([None] * (len(current["headers"]) - len(normalized_row)))

            record = {
                header: self.normalize_cell_value(normalized_row[idx])
                for idx, header in enumerate(current["headers"])
            }
            if self._record_has_content(record):
                current["rows"].append(record)
                variety = self._extract_variety(record)
                if variety:
                    current["varieties"].add(variety)

        if current and current["rows"]:
            blocks.append(current)

        if not blocks:
            non_empty_rows = [row for row in rows if self._row_has_content(row)]
            if not non_empty_rows:
                return []
            header_idx = self.detect_header_row(non_empty_rows)
            headers = self._build_headers_from_row(non_empty_rows[header_idx])
            fallback_rows: list[dict[str, Any]] = []
            for row in non_empty_rows[header_idx + 1 :]:
                normalized_row = list(row)
                if len(normalized_row) < len(headers):
                    normalized_row.extend([None] * (len(headers) - len(normalized_row)))
                record = {header: self.normalize_cell_value(normalized_row[idx]) for idx, header in enumerate(headers)}
                if self._record_has_content(record):
                    fallback_rows.append(record)
            blocks.append({"headers": headers, "header_row_index": header_idx, "rows": fallback_rows, "varieties": set()})
        return blocks

    def _is_header_candidate(self, row: list[Any]) -> bool:
        texts = [self._normalize_header_token(self._cell_to_text(value)) for value in row]
        non_empty = [t for t in texts if t]
        if len(non_empty) < 2:
            return False
        keyword_hits = sum(1 for t in non_empty if t in self.HEADER_KEYWORDS)
        if keyword_hits >= 2:
            return True
        unique_text = set(non_empty)
        return len(unique_text.intersection(self.HEADER_KEYWORDS)) >= 2

    def _build_headers_from_row(self, row: list[Any]) -> list[str]:
        used_headers: dict[str, int] = {}
        headers = [self.clean_header(value, idx, used_headers) for idx, value in enumerate(row)]
        return headers if headers else ["Columna_1"]

    def _headers_compatible(self, base_headers: list[str], other_headers: list[str]) -> bool:
        base_norm = {self._normalize_header_token(header): header for header in base_headers}
        other_norm = {self._normalize_header_token(header): header for header in other_headers}
        shared = set(base_norm).intersection(other_norm)
        if len(shared) >= 2:
            return True
        expected = {"socio", "nombre", "variedad", "neto", "importe", "precio"}
        return len(shared.intersection(expected)) >= 2

    def _remap_rows(self, rows: list[dict[str, Any]], source_headers: list[str], target_headers: list[str]) -> list[dict[str, Any]]:
        source_map = {self._normalize_header_token(header): header for header in source_headers}
        remapped: list[dict[str, Any]] = []
        for row in rows:
            mapped_row: dict[str, Any] = {}
            for target in target_headers:
                source_key = source_map.get(self._normalize_header_token(target))
                mapped_row[target] = row.get(source_key) if source_key else None
            if self._record_has_content(mapped_row):
                remapped.append(mapped_row)
        return remapped

    def _extract_variety(self, record: dict[str, Any]) -> str:
        for key, value in record.items():
            normalized = self._normalize_header_token(key)
            if normalized == "variedad":
                return str(value or "").strip().upper()
        return ""

    def _sum_numeric_column(self, rows: list[dict[str, Any]], headers: list[str], semantic_header: str) -> float:
        target = None
        for header in headers:
            if self._normalize_header_token(header) == semantic_header:
                target = header
                break
        if not target:
            return 0.0
        total = 0.0
        for row in rows:
            value = row.get(target)
            if isinstance(value, (int, float)):
                total += float(value)
        return total

    def _normalize_header_token(self, value: str) -> str:
        return "".join(ch.lower() for ch in value.strip() if ch.isalnum())

    def _read_csv(self, path: str, encoding: str, forced_delimiter: str | None) -> list[list[str]]:
        with open(path, "r", encoding=encoding, newline="") as file:
            sample = file.read(4096)
            file.seek(0)
            if forced_delimiter:
                reader = csv.reader(file, delimiter=forced_delimiter)
            else:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,")
                reader = csv.reader(file, dialect)
            return [row for row in reader]

    def _parse_numeric_string(self, value: str) -> int | float | None:
        text = value.replace(" ", "")
        if not text:
            return None

        comma_count = text.count(",")
        dot_count = text.count(".")

        normalized = text
        if comma_count > 0 and dot_count > 0:
            if text.rfind(",") > text.rfind("."):
                normalized = text.replace(".", "").replace(",", ".")
            else:
                normalized = text.replace(",", "")
        elif comma_count > 0:
            if comma_count == 1:
                normalized = text.replace(",", ".")
            else:
                normalized = text.replace(",", "")

        try:
            number = float(normalized)
        except ValueError:
            return None

        if number.is_integer() and all(sep not in value for sep in (",", ".")):
            return int(number)
        return number

    def _looks_numeric(self, value: str) -> bool:
        return self._parse_numeric_string(value) is not None

    def _cell_to_text(self, value: Any) -> str:
        return "" if value is None else str(value)

    def _row_has_content(self, row: list[Any]) -> bool:
        return any(self._cell_to_text(value).strip() for value in row)

    def _record_has_content(self, row: dict[str, Any]) -> bool:
        for value in row.values():
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return True
        return False
